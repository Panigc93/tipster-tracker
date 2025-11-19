#!/usr/bin/env python3
"""Comparar estructura de Excel generado vs template original"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

original = '/home/cgarciap/Escritorio/tipster-tracker/template-picks-and-tipsters.xlsx'
generado = '/home/cgarciap/Escritorio/tipster-tracker-template-2025-11-19.xlsx'

print("=" * 100)
print("COMPARACIÓN: Template Original vs Template Generado")
print("=" * 100)

wb_orig = load_workbook(original)
wb_gen = load_workbook(generado)

# Comparar sheets
print("\n📋 SHEETS:")
print(f"  Original: {wb_orig.sheetnames}")
print(f"  Generado: {wb_gen.sheetnames}")

sheets_orig = set(wb_orig.sheetnames)
sheets_gen = set(wb_gen.sheetnames)
if sheets_orig == sheets_gen:
    print("  ✅ Mismo número y nombres de sheets")
else:
    print(f"  ⚠️  Diferencias en sheets:")
    missing = sheets_orig - sheets_gen
    extra = sheets_gen - sheets_orig
    if missing:
        print(f"     Faltan: {missing}")
    if extra:
        print(f"     Extras: {extra}")

# Comparar estructura de cada sheet
for sheet_name in ['Lanzadas Tipster', 'Realizadas']:
    if sheet_name not in wb_gen.sheetnames:
        continue
        
    print(f"\n{'=' * 100}")
    print(f"📄 SHEET: {sheet_name}")
    print('=' * 100)
    
    ws_orig = wb_orig[sheet_name]
    ws_gen = wb_gen[sheet_name]
    
    # Comparar headers de fila 6
    print("\n🔍 HEADERS (Fila 6):")
    for col_idx in range(1, min(ws_orig.max_column, ws_gen.max_column) + 1):
        col_letter = get_column_letter(col_idx)
        header_orig = ws_orig.cell(6, col_idx).value
        header_gen = ws_gen.cell(6, col_idx).value
        
        if header_orig or header_gen:
            match = "✅" if header_orig == header_gen else "❌"
            print(f"  {col_letter}: {match}")
            if header_orig != header_gen:
                print(f"     Original: {header_orig}")
                print(f"     Generado: {header_gen}")
    
    # Comparar fórmulas de fila 2
    print(f"\n🧮 FÓRMULAS (Fila 2):")
    for col_idx in range(1, min(11, ws_orig.max_column + 1)):
        col_letter = get_column_letter(col_idx)
        cell_orig = ws_orig.cell(2, col_idx)
        cell_gen = ws_gen.cell(2, col_idx)
        
        formula_orig = cell_orig.value if isinstance(cell_orig.value, str) and cell_orig.value.startswith('=') else None
        formula_gen = cell_gen.value if isinstance(cell_gen.value, str) and cell_gen.value.startswith('=') else None
        
        if formula_orig or formula_gen:
            match = "✅" if formula_orig == formula_gen else "❌"
            print(f"  {col_letter}2: {match}")
            if formula_orig != formula_gen:
                print(f"     Original: {formula_orig}")
                print(f"     Generado: {formula_gen}")

print("\n" + "=" * 100)
print("✅ COMPARACIÓN COMPLETA")
print("=" * 100)
print("\n💡 Abre el archivo en LibreOffice/Excel para verificar visualmente:")
print(f"   {generado}")
