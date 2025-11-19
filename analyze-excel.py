#!/usr/bin/env python3
"""Script para analizar la estructura del template de Excel"""

import pandas as pd
from openpyxl import load_workbook

# Cargar el archivo Excel
file_path = '/home/cgarciap/Escritorio/tipster-tracker/template-picks-and-tipsters.xlsx'
print("=" * 80)
print("ANÁLISIS DEL TEMPLATE EXCEL")
print("=" * 80)

# 1. Listar todas las sheets con pandas
xl_file = pd.ExcelFile(file_path)
print(f"\n📋 SHEETS DISPONIBLES ({len(xl_file.sheet_names)}):")
for i, sheet in enumerate(xl_file.sheet_names, 1):
    print(f"  {i}. {sheet}")

# 2. Analizar cada sheet
wb = load_workbook(file_path)

for sheet_name in xl_file.sheet_names:
    print(f"\n{'=' * 80}")
    print(f"📄 SHEET: '{sheet_name}'")
    print('=' * 80)
    
    ws = wb[sheet_name]
    print(f"  Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
    
    # Leer con pandas para mejor visualización
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=10)
    
    # Mostrar primeras 5 filas
    print("\n  PRIMERAS 5 FILAS (preview):")
    print(df.head(5).to_string(index=False, max_colwidth=20))
    
    # Header detection (fila 1)
    print("\n  HEADERS (Fila 1):")
    headers_found = False
    for col_idx in range(1, min(ws.max_column + 1, 20)):
        cell = ws.cell(1, col_idx)
        if cell.value:
            headers_found = True
            col_letter = cell.column_letter
            print(f"    {col_letter}1: {cell.value}")
    if not headers_found:
        print("    (No se encontraron headers)")
    
    # Detectar fórmulas en primeras 10 filas
    print("\n  FÓRMULAS DETECTADAS (primeras 10 filas):")
    has_formulas = False
    for row_idx in range(1, min(11, ws.max_row + 1)):
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                has_formulas = True
                col_letter = cell.column_letter
                formula = cell.value[:80] + ("..." if len(cell.value) > 80 else "")
                print(f"    {col_letter}{row_idx}: {formula}")
    if not has_formulas:
        print("    (No se detectaron fórmulas en las primeras 10 filas)")
    
    # Detectar validaciones de datos (dropdowns)
    print("\n  VALIDACIONES DE DATOS (dropdowns):")
    has_validations = False
    if ws.data_validations and ws.data_validations.dataValidation:
        for dv in ws.data_validations.dataValidation:
            if dv.formula1:
                has_validations = True
                print(f"    Rangos: {dv.sqref}")
                print(f"    Tipo: {dv.type}")
                print(f"    Fórmula: {dv.formula1}")
    if not has_validations:
        print("    (No se detectaron validaciones de datos)")

print(f"\n{'=' * 80}")
print("✅ ANÁLISIS COMPLETO")
print('=' * 80)
