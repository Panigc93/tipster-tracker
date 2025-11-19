#!/usr/bin/env python3
"""Análisis detallado de las columnas de entrada"""

import pandas as pd
from openpyxl import load_workbook

file_path = '/home/cgarciap/Escritorio/tipster-tracker/template-picks-and-tipsters.xlsx'

print("=" * 80)
print("ANÁLISIS DETALLADO DE COLUMNAS DE ENTRADA")
print("=" * 80)

# Analizar sheets de entrada
input_sheets = ['Realizadas', 'Lanzadas Tipster']

for sheet_name in input_sheets:
    print(f"\n{'=' * 80}")
    print(f"📋 SHEET: '{sheet_name}'")
    print('=' * 80)
    
    # Leer primeras filas
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=20)
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    print("\n🔍 ANÁLISIS DE ESTRUCTURA:")
    print(f"  - Fila 1: Headers/Estadísticas generales")
    print(f"  - Fila 2: Fórmulas auto-calculadas")
    print(f"  - Fila 3-6: Instrucciones/plantilla")
    print(f"  - Fila 7+: DATOS DE ENTRADA (picks)")
    
    print("\n📊 COLUMNAS DE ENTRADA (a partir de fila 7):")
    
    # Mapeo manual basado en el análisis visual
    columns_mapping = {
        'A': 'LIVE/PRE',
        'B': 'Tipster',
        'C': 'Stake (UDS)',
        'D': 'Odds (Cuota)',
        'E': 'Resultado (W/L/V)',
        'F': 'Beneficio (auto-calc)',
        'G': 'Unidades Apostadas (auto-calc)',
        'H': 'Beneficio Real (auto-calc)',
        # Continuar con otras columnas relevantes
    }
    
    print("\n  COLUMNAS PRINCIPALES:")
    for col_letter, col_name in columns_mapping.items():
        cell_value = ws[f"{col_letter}7"].value if ws[f"{col_letter}7"].value else "(vacío)"
        print(f"    {col_letter}: {col_name}")
        print(f"       Ejemplo fila 7: {cell_value}")
    
    # Detectar todas las columnas con headers
    print("\n  TODAS LAS COLUMNAS DETECTADAS:")
    for col_idx in range(1, min(ws.max_column + 1, 30)):
        col_letter = ws.cell(1, col_idx).column_letter
        header = ws.cell(1, col_idx).value
        
        if header:
            print(f"    {col_letter}1: {header}")
            
            # Ver si tiene fórmula o es entrada manual
            cell_7 = ws.cell(7, col_idx)
            if cell_7.value and isinstance(cell_7.value, str) and cell_7.value.startswith('='):
                print(f"       ⚙️  AUTO-CALCULADA: {cell_7.value[:60]}...")
            else:
                print(f"       ✍️  ENTRADA MANUAL")

print(f"\n{'=' * 80}")
print("🎯 RESUMEN PARA MAPEO:")
print('=' * 80)
print("""
MAPEO PROPUESTO (Excel → App):

Sheet "Realizadas" (UserFollows):
  - Columna A: pickType (LIVE/PRE/Combinado)
  - Columna B: tipsterName (dropdown)
  - Columna C: userStake (manual)
  - Columna D: userOdds (manual)
  - Columna E: userResult (W=Ganada, L=Perdida, V=Void)
  - Otras columnas: sport, bookmaker, match, betType, etc.

Sheet "Lanzadas Tipster" (Picks):
  - Columna A: pickType (LIVE/PRE/Combinado)
  - Columna B: tipsterName (dropdown)
  - Columna C: stake (manual)
  - Columna D: odds (manual)
  - Columna E: result (W=Ganada, L=Perdida, V=Void)
  - Otras columnas: sport, bookmaker, match, betType, etc.

Sheet "Base datos" (Reference Data):
  - Columna A: Bookmakers
  - Columna C: Canales (Blogabet, Telegram, etc.)
  - Columna E: Deportes
  - Columna G: LIVE/PRE options
  - Columna I: Combinada (Si/No)
""")
