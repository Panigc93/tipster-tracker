#!/usr/bin/env python3
"""
Script para añadir estilos y fórmulas al Excel generado por la app React.

Funcionalidades:
- Estilos: Fondo azul en celdas APUESTA fusionadas (M6:O6)
- Fórmulas: Añade fórmulas a las filas de datos (F, G, H) en filas 7-10

Uso:
    python3 scripts/add-excel-styles.py <archivo.xlsx>
    
Ejemplo:
    python3 scripts/add-excel-styles.py ~/Escritorio/export-picks.xlsx
"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    print("❌ Error: openpyxl no está instalado")
    print("Instalar con: pip3 install openpyxl")
    sys.exit(1)


def add_styles_and_formulas_to_excel(filepath: str) -> None:
    """
    Añade estilos, fórmulas y dropdowns al Excel:
    - Fondo azul en celdas APUESTA fusionadas (M6:O6)
    - Texto blanco y negrita
    - Fórmulas en columnas F, G, H para filas 7-10 (plantilla)
    - Data validations (dropdowns) en todas las columnas necesarias
    """
    path = Path(filepath)
    
    if not path.exists():
        print(f"❌ Error: Archivo no encontrado: {filepath}")
        sys.exit(1)
    
    print(f"📂 Abriendo: {filepath}")
    
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        
        wb = load_workbook(filepath)
        
        # Estilos (Arial en todas las fuentes)
        blue_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        white_font = Font(name='Arial', color='FFFFFF', bold=True, size=11)
        white_font_large = Font(name='Arial', color='FFFFFF', bold=True, size=14)
        black_font_large = Font(name='Arial', color='000000', bold=True, size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        sheets_to_process = ['Lanzadas Tipster', 'Realizadas']
        
        for sheet_name in sheets_to_process:
            if sheet_name not in wb.sheetnames:
                print(f"⚠️  Hoja '{sheet_name}' no encontrada, saltando...")
                continue
            
            ws = wb[sheet_name]
            
            # ========================================
            # 1. CELDAS GRANDES FUSIONADAS (A3:E5 y F3:T5)
            # ========================================
            # A3:E5 - "PLANTILLA EII" con fondo negro y letras blancas
            ws.merge_cells('A3:E5')
            ws['A3'].value = 'PLANTILLA EII'
            ws['A3'].fill = black_fill
            ws['A3'].font = white_font_large
            ws['A3'].alignment = center_alignment
            
            # F3:T5 - "Picks que lanza el tipster" con fondo negro y letras blancas
            ws.merge_cells('F3:T5')
            ws['F3'].value = 'Rellenar con los picks que lanza el tipster'
            ws['F3'].fill = black_fill
            ws['F3'].font = white_font_large
            ws['F3'].alignment = center_alignment
            
            # ========================================
            # 2. ICONOS CENTRADOS EN FILA 1 (✅❌🔵)
            # ========================================
            # B1: ✅ con fondo verde (icono de Win)
            ws['B1'].alignment = center_alignment
            ws['B1'].fill = green_fill
            
            # C1: ❌ centrado
            ws['C1'].alignment = center_alignment
            
            # D1: 🔵 centrado
            ws['D1'].alignment = center_alignment
            
            # ========================================
            # 3. FILA 6 (Headers) - Altura aumentada
            # ========================================
            ws.row_dimensions[6].height = 25
            
            # Headers con fondo azul, texto negro mediano y centrado
            black_font_medium = Font(name='Arial', color='000000', bold=True, size=11)
            headers_blue = ['B6', 'C6', 'D6', 'E6', 'M6']  # TIPSTER, STAKE, CUOTA, W/L/V, APUESTA
            for cell_ref in headers_blue:
                cell = ws[cell_ref]
                cell.fill = blue_fill
                cell.font = black_font_medium
                cell.alignment = center_alignment
            
            # ========================================
            # 4. FORMATO PORCENTAJE A YIELD (G2)
            # ========================================
            ws['G2'].number_format = '0.00%'
            
            # ========================================
            # 4.5. FORMATO CONDICIONAL EN FILA 2 (E2, F2, G2, H2)
            # ========================================
            from openpyxl.formatting.rule import CellIsRule
            
            light_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            
            # E2, F2, G2, H2: Rojo si < 0, Verde si > 0
            for col_letter in ['E', 'F', 'G', 'H']:
                cell_ref = f'{col_letter}2'
                
                # Regla: Rojo si negativo
                ws.conditional_formatting.add(
                    cell_ref,
                    CellIsRule(operator='lessThan', formula=['0'], fill=light_red_fill)
                )
                
                # Regla: Verde si positivo
                ws.conditional_formatting.add(
                    cell_ref,
                    CellIsRule(operator='greaterThan', formula=['0'], fill=light_green_fill)
                )
            
            # ========================================
            # 5. FONDO AZUL CLARITO EN COLUMNAS B, C, D, E (desde fila 7)
            # ========================================
            from openpyxl.styles import Border, Side
            
            light_blue_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Aplicar estilos solo a las primeras 50 filas (template)
            # El usuario puede copiar el formato después si necesita más filas
            max_row_for_styles = 50
            
            # Aplicar fondo azul clarito a columnas B, C, D, E desde fila 7
            # El formato condicional de la columna E lo sobrescribirá cuando haya W/L
            for row in range(7, max_row_for_styles + 1):
                for col in ['B', 'C', 'D', 'E']:
                    cell = ws[f'{col}{row}']
                    cell.fill = light_blue_fill
                    cell.border = thin_border
                    cell.alignment = center_alignment
            
            # Columna E: formato condicional según W/L/V (filas 7-50)
            # Rojo si contiene "L", Verde si contiene "W", mantiene azul si V o vacío
            ws.conditional_formatting.add(
                f'E7:E{max_row_for_styles}',
                CellIsRule(operator='equal', formula=['"L"'], fill=light_red_fill)
            )
            ws.conditional_formatting.add(
                f'E7:E{max_row_for_styles}',
                CellIsRule(operator='equal', formula=['"W"'], fill=light_green_fill)
            )
            
            # Aplicar bordes a todas las celdas que se pueden rellenar (A, F-T desde fila 7)
            for row in range(7, max_row_for_styles + 1):
                for col_letter in ['A', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
                    cell = ws[f'{col_letter}{row}']
                    cell.border = thin_border
                    cell.alignment = center_alignment
            
            print(f"✅ Estilos aplicados a '{sheet_name}' - Celdas fusionadas, iconos, headers, yield %, fondos y bordes")
            
            # ========================================
            # 6. AÑADIR FÓRMULAS EN FILAS DE DATOS (7-10)
            # ========================================
            data_rows = [7, 8, 9, 10]
            
            for row in data_rows:
                # F: Resultado unidades
                # =IF(E7="L",-C7,IF(E7="W",C7*(D7-1),IF(E7="HW",(C7/2)*(D7-1),IF(E7="HL",-C7/2,0))))
                ws[f'F{row}'].value = f'=IF(E{row}="L",-C{row},IF(E{row}="W",C{row}*(D{row}-1),IF(E{row}="HW",(C{row}/2)*(D{row}-1),IF(E{row}="HL",-C{row}/2,0))))'
                
                # G: CANTIDAD (con VLOOKUP)
                # =IFERROR(($I$2/VLOOKUP(B7,Tipster_Picks_Dashboard!$A$3:$W$76,2,FALSE))*C7,"")
                ws[f'G{row}'].value = f'=IFERROR(($I$2/VLOOKUP(B{row},Tipster_Picks_Dashboard!$A$3:$W$76,2,FALSE))*C{row},"")'
                
                # H: Resultado euros
                # =IF(E7="w",G7*D7-G7,IF(E7="L",-G7,0))
                ws[f'H{row}'].value = f'=IF(E{row}="w",G{row}*D{row}-G{row},IF(E{row}="L",-G{row},0))'
            
            print(f"✅ Fórmulas añadidas a '{sheet_name}' - Filas 7-10 (F, G, H)")
            
            # ========================================
            # 3. AÑADIR DROPDOWNS (DATA VALIDATIONS)
            # ========================================
            
            # Determinar rangos según la hoja (1853 para Lanzadas, 2001 para Realizadas)
            max_row = 1853 if sheet_name == 'Lanzadas Tipster' else 2001
            
            # A: LIVE-PRE (PRE/LIVE/Combinado)
            dv_livepre = DataValidation(
                type="list",
                formula1="'Base datos'!$G$2:$G$4",
                allow_blank=True,
                showDropDown=False  # False muestra la flecha (counter-intuitive Excel quirk)
            )
            dv_livepre.error = 'Selecciona: PRE, LIVE o Combinado'
            dv_livepre.errorTitle = 'Valor no válido'
            ws.add_data_validation(dv_livepre)
            dv_livepre.add(f'A7:A{max_row}')
            
            # B: TIPSTER (dinámico desde dashboard correspondiente)
            # Realizadas → Mis_Picks_Dashboard
            # Lanzadas Tipster → Tipster_Picks_Dashboard
            tipster_formula = "Mis_Picks_Dashboard!$A$3:$A$100" if sheet_name == 'Realizadas' else "Tipster_Picks_Dashboard!$A$3:$A$100"
            
            dv_tipster = DataValidation(
                type="list",
                formula1=tipster_formula,
                allow_blank=True,
                showDropDown=False  # False muestra la flecha
            )
            dv_tipster.error = 'Selecciona un tipster de la lista'
            dv_tipster.errorTitle = 'Tipster no válido'
            ws.add_data_validation(dv_tipster)
            dv_tipster.add(f'B7:B{max_row}')
            
            # E: W/L/V/HW/HL (resultados)
            dv_result = DataValidation(
                type="list",
                formula1='"W,L,V,HW,HL"',
                allow_blank=True,
                showDropDown=False  # False muestra la flecha
            )
            dv_result.error = 'Selecciona: W (Win), L (Loss), V (Void), HW (Half Win), HL (Half Loss)'
            dv_result.errorTitle = 'Resultado no válido'
            ws.add_data_validation(dv_result)
            dv_result.add(f'E7:E{max_row}')
            
            # Dropdowns que difieren según la hoja
            if sheet_name == 'Lanzadas Tipster':
                # P: COMBINADA (Si/No)
                dv_combinada = DataValidation(
                    type="list",
                    formula1="'Base datos'!$I$2:$I$4",  # Columna I, filas 2-4 (Si, No, Sí)
                    allow_blank=True,
                    showDropDown=False  # False muestra la flecha
                )
                dv_combinada.error = 'Selecciona: Si o No'
                dv_combinada.errorTitle = 'Valor no válido'
                ws.add_data_validation(dv_combinada)
                dv_combinada.add(f'P7:P{max_row}')
                
                # Q: DEPORTE
                dv_deporte = DataValidation(
                    type="list",
                    formula1="'Base datos'!$E$2:$E$20",
                    allow_blank=True,
                    showDropDown=False  # False muestra la flecha
                )
                dv_deporte.error = 'Selecciona un deporte de la lista'
                dv_deporte.errorTitle = 'Deporte no válido'
                ws.add_data_validation(dv_deporte)
                dv_deporte.add(f'Q7:Q{max_row}')
                
                # R: Plataforma envio pick
                dv_plataforma = DataValidation(
                    type="list",
                    formula1="'Base datos'!$C$2:$C$20",
                    allow_blank=True,
                    showDropDown=False  # False muestra la flecha
                )
                dv_plataforma.error = 'Selecciona una plataforma de la lista'
                dv_plataforma.errorTitle = 'Plataforma no válida'
                ws.add_data_validation(dv_plataforma)
                dv_plataforma.add(f'R7:R{max_row}')
                
                # S: BOOKIE recomendado
                dv_bookie = DataValidation(
                    type="list",
                    formula1="'Base datos'!$A$2:$A$30",
                    allow_blank=True,
                    showDropDown=False  # False muestra la flecha
                )
                dv_bookie.error = 'Selecciona un bookmaker de la lista'
                dv_bookie.errorTitle = 'Bookmaker no válido'
                ws.add_data_validation(dv_bookie)
                dv_bookie.add(f'S7:S{max_row}')
                
            else:  # Realizadas
                # Q: COMBINADA (columna desplazada por Comentarios en P)
                dv_combinada = DataValidation(
                    type="list",
                    formula1="'Base datos'!$I$2:$I$4",  # Columna I, filas 2-4 (Si, No, Sí)
                    allow_blank=True,
                    showDropDown=False  # False muestra la flecha
                )
                dv_combinada.error = 'Selecciona: Si o No'
                dv_combinada.errorTitle = 'Valor no válido'
                ws.add_data_validation(dv_combinada)
                dv_combinada.add(f'Q7:Q{max_row}')
                
                # R: DEPORTE
                dv_deporte = DataValidation(
                    type="list",
                    formula1="'Base datos'!$E$2:$E$20",
                    allow_blank=True,
                    showDropDown=False  # False muestra la flecha
                )
                dv_deporte.error = 'Selecciona un deporte de la lista'
                dv_deporte.errorTitle = 'Deporte no válido'
                ws.add_data_validation(dv_deporte)
                dv_deporte.add(f'R7:R{max_row}')
                
                # S: Plataforma envio pick
                dv_plataforma = DataValidation(
                    type="list",
                    formula1="'Base datos'!$C$2:$C$20",
                    allow_blank=True,
                    showDropDown=False  # False muestra la flecha
                )
                dv_plataforma.error = 'Selecciona una plataforma de la lista'
                dv_plataforma.errorTitle = 'Plataforma no válida'
                ws.add_data_validation(dv_plataforma)
                dv_plataforma.add(f'S7:S{max_row}')
                
                # T: BOOKIE (sin "recomendado")
                dv_bookie = DataValidation(
                    type="list",
                    formula1="'Base datos'!$A$2:$A$30",
                    allow_blank=True,
                    showDropDown=False  # False muestra la flecha
                )
                dv_bookie.error = 'Selecciona un bookmaker de la lista'
                dv_bookie.errorTitle = 'Bookmaker no válido'
                ws.add_data_validation(dv_bookie)
                dv_bookie.add(f'T7:T{max_row}')
            
            print(f"✅ Dropdowns añadidos a '{sheet_name}' - Columnas A, B, E, P/Q, Q/R, R/S, S/T")
        
        # ========================================
        # ESTILAR HOJAS DASHBOARD
        # ========================================
        style_dashboard_sheets(wb)
        
        # ========================================
        # COPIAR FÓRMULAS EN DASHBOARDS (FILAS 4-100)
        # ========================================
        copy_dashboard_formulas(wb)
        
        # ========================================
        # APLICAR FUENTE ARIAL A TODO EL EXCEL
        # ========================================
        apply_arial_to_all_sheets(wb)
        
        # Guardar
        backup_path = path.with_suffix('.backup.xlsx')
        print(f"💾 Creando backup: {backup_path.name}")
        wb.save(backup_path)
        
        print(f"💾 Guardando cambios: {path.name}")
        wb.save(filepath)
        
        print("\n✅ ¡Estilos y fórmulas aplicados exitosamente!")
        print(f"📁 Archivo actualizado: {filepath}")
        print(f"📁 Backup guardado: {backup_path}")
        
    except Exception as e:
        print(f"❌ Error al procesar el archivo: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def style_dashboard_sheets(wb):
    """
    Aplica estilos a las hojas dashboard (Mis_Picks_Dashboard y Tipster_Picks_Dashboard)
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Border, Side
    
    # Estilos
    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    light_blue_fill = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    light_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    white_font = Font(name='Arial', color='FFFFFF', bold=True, size=14)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    dashboard_sheets = ['Mis_Picks_Dashboard', 'Tipster_Picks_Dashboard']
    
    for sheet_name in dashboard_sheets:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        print(f"\n🎨 Estilando hoja: {sheet_name}")
        
        # Fuente pequeña para dashboard
        small_font = Font(name='Arial', size=9)
        
        # ========================================
        # 1. TÍTULOS DE LA TABLA (FILA 2)
        # ========================================
        # Columna A (TIPSTER): Azul clarito
        ws['A2'].fill = light_blue_fill
        ws['A2'].border = thin_border
        ws['A2'].alignment = center_alignment
        ws['A2'].font = small_font
        
        # Columnas B-M (resto de títulos): Amarillo clarito
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            cell = ws[f'{col_letter}2']
            cell.fill = yellow_fill
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.font = small_font
        
        # Columnas N-AC (16 deportes): Amarillo clarito también
        for col_letter in ['N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']:
            cell = ws[f'{col_letter}2']
            cell.fill = yellow_fill
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.font = small_font
        
        # ========================================
        # 2. CELDA N1 "% Aciertos Segun deporte" (N1:AC1)
        # ========================================
        # Fusionar N1:AC1 (16 deportes)
        ws.merge_cells('N1:AC1')
        ws['N1'].value = '% Aciertos Segun deporte'
        ws['N1'].fill = black_fill
        ws['N1'].font = white_font
        ws['N1'].alignment = center_alignment
        
        # ========================================
        # 3. DATOS (FILA 3 EN ADELANTE) - Bordes negros y fuente pequeña
        # ========================================
        # Aplicar bordes y fuente pequeña a las primeras 100 filas de datos (A-AC: 29 columnas)
        for row in range(3, 103):
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC']:
                cell = ws[f'{col_letter}{row}']
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.font = small_font
        
        # Columna A (TIPSTER) desde fila 3: Azul clarito
        for row in range(3, 103):
            ws[f'A{row}'].fill = light_blue_fill
        
        # ========================================
        # 3.5. ANCHOS DE COLUMNA MÁS ESTRECHOS
        # ========================================
        ws.column_dimensions['A'].width = 12  # TIPSTER
        ws.column_dimensions['B'].width = 8   # Juega a Unidades
        ws.column_dimensions['C'].width = 8   # Benficio UDS
        ws.column_dimensions['D'].width = 8   # Beneficio
        ws.column_dimensions['E'].width = 8   # Apostado
        ws.column_dimensions['F'].width = 6   # YIELD
        ws.column_dimensions['G'].width = 8   # Tips totales
        ws.column_dimensions['H'].width = 6   # Tips W
        ws.column_dimensions['I'].width = 6   # Tips L
        ws.column_dimensions['J'].width = 6   # Tips V
        ws.column_dimensions['K'].width = 8   # % Tips Aciertados
        ws.column_dimensions['L'].width = 8   # % Aciertos Live
        ws.column_dimensions['M'].width = 8   # % Aciertos PRE
        # 16 deportes (N-AC):
        ws.column_dimensions['N'].width = 7   # Badminton
        ws.column_dimensions['O'].width = 7   # Baloncesto
        ws.column_dimensions['P'].width = 7   # Balonmano
        ws.column_dimensions['Q'].width = 7   # Beisbol
        ws.column_dimensions['R'].width = 6   # Boxeo
        ws.column_dimensions['S'].width = 7   # Ciclismo
        ws.column_dimensions['T'].width = 7   # Esports
        ws.column_dimensions['U'].width = 7   # Fútbol
        ws.column_dimensions['V'].width = 8   # Fútbol Americano
        ws.column_dimensions['W'].width = 6   # Golf
        ws.column_dimensions['X'].width = 7   # Hockey
        ws.column_dimensions['Y'].width = 6   # MMA
        ws.column_dimensions['Z'].width = 7   # Tenis
        ws.column_dimensions['AA'].width = 7  # Tenis Mesa
        ws.column_dimensions['AB'].width = 7  # Voleibol
        ws.column_dimensions['AC'].width = 7  # (extra si se añade)
        
        # ========================================
        # 4. FORMATO CONDICIONAL (Columnas C, D, F desde fila 3)
        # ========================================
        # C (Benficio UDS), D (Beneficio), F (YIELD): Rojo si < 0, Verde si > 0
        for col_letter in ['C', 'D', 'F']:
            cell_range = f'{col_letter}3:{col_letter}100'
            
            # Regla: Rojo si negativo
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='lessThan', formula=['0'], fill=light_red_fill)
            )
            
            # Regla: Verde si positivo
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='greaterThan', formula=['0'], fill=light_green_fill)
            )
        
        print(f"✅ Estilos aplicados a '{sheet_name}' - Títulos, bordes, formato condicional")


def copy_dashboard_formulas(wb):
    """
    Copia las fórmulas de la fila 3 (ejemplo "Manolo") a las filas 4-100
    para que el usuario pueda escribir nuevos tipsters y las fórmulas se calculen automáticamente.
    """
    from openpyxl.utils import get_column_letter
    from copy import copy
    
    dashboard_sheets = ['Mis_Picks_Dashboard', 'Tipster_Picks_Dashboard']
    
    for sheet_name in dashboard_sheets:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        print(f"\n📋 Copiando fórmulas en: {sheet_name}")
        
        # Copiar fórmulas de fila 3 a filas 4-100
        # Columnas C-T (índices 3-20) tienen fórmulas
        for row_num in range(4, 101):  # Filas 4 a 100
            for col_num in range(3, 21):  # Columnas C(3) a T(20)
                col_letter = get_column_letter(col_num)
                
                # Celda fuente (fila 3) y celda destino (fila actual)
                source_cell = ws[f'{col_letter}3']
                target_cell = ws[f'{col_letter}{row_num}']
                
                # Copiar fórmula si existe
                if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                    # Reemplazar referencias de fila 3 por fila actual
                    formula = source_cell.value
                    # Reemplazar A3 por A{row_num}, etc.
                    formula = formula.replace('A3', f'A{row_num}')
                    formula = formula.replace('H3', f'H{row_num}')
                    formula = formula.replace('D3', f'D{row_num}')
                    formula = formula.replace('E3', f'E{row_num}')
                    formula = formula.replace('G3', f'G{row_num}')
                    
                    target_cell.value = formula
                    
                    # Copiar estilos
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.protection = copy(source_cell.protection)
                        target_cell.alignment = copy(source_cell.alignment)
        
        print(f"✅ Fórmulas copiadas a 97 filas (4-100) en '{sheet_name}'")


def apply_arial_to_all_sheets(wb):
    """
    Aplica fuente Arial a TODAS las celdas de TODAS las hojas del Excel.
    """
    from openpyxl.styles import Font
    from copy import copy
    
    arial_font = Font(name='Arial')
    
    print("\n🔤 Aplicando fuente Arial a todo el Excel...")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"   📄 {sheet_name}...")
        
        # Iterar sobre todas las filas y columnas usadas
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:  # Solo celdas con contenido
                    # Copiar font existente y cambiar solo el name
                    if cell.font:
                        new_font = copy(cell.font)
                        new_font = Font(
                            name='Arial',
                            size=new_font.size,
                            bold=new_font.bold,
                            italic=new_font.italic,
                            vertAlign=new_font.vertAlign,
                            underline=new_font.underline,
                            strike=new_font.strike,
                            color=new_font.color
                        )
                        cell.font = new_font
                    else:
                        cell.font = arial_font
    
    print("✅ Fuente Arial aplicada a todas las hojas")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("❌ Error: Falta el archivo Excel")
        print("\nUso:")
        print("    python3 scripts/add-excel-styles.py <archivo.xlsx>")
        print("\nEjemplo:")
        print("    python3 scripts/add-excel-styles.py ~/Escritorio/EXCEL-V3-FINAL.xlsx")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    add_styles_and_formulas_to_excel(excel_file)
