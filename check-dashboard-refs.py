#!/usr/bin/env python3
"""
Script para verificar a qué dashboard apuntan las fórmulas VLOOKUP
en las hojas Realizadas y Lanzadas Tipster
"""

import openpyxl
import sys

def check_dashboard_references(filename):
    print(f"\n📊 Analizando: {filename}")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(filename, data_only=False)
    
    sheets_to_check = ['Realizadas', 'Lanzadas Tipster']
    
    for sheet_name in sheets_to_check:
        if sheet_name not in wb.sheetnames:
            print(f"\n⚠️  Sheet '{sheet_name}' no encontrada")
            continue
            
        ws = wb[sheet_name]
        print(f"\n📋 SHEET: '{sheet_name}'")
        print("-" * 80)
        
        # Revisar fila 7, columna G (donde está el VLOOKUP)
        cell_g7 = ws['G7']
        
        if cell_g7.value and isinstance(cell_g7.value, str) and 'VLOOKUP' in cell_g7.value:
            formula = cell_g7.value
            print(f"\n✅ Fórmula encontrada en G7:")
            print(f"   {formula}")
            
            # Extraer el nombre del dashboard
            if 'Mis_Picks_Dashboard' in formula:
                print(f"\n🎯 Referencia: Mis_Picks_Dashboard")
            elif 'Tipster_Picks_Dashboard' in formula:
                print(f"\n🎯 Referencia: Tipster_Picks_Dashboard")
            else:
                print(f"\n⚠️  Dashboard no identificado")
        else:
            print(f"\n❌ No se encontró VLOOKUP en G7")
            print(f"   Valor de G7: {cell_g7.value}")
    
    wb.close()
    print("\n" + "=" * 80 + "\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        check_dashboard_references(sys.argv[1])
    else:
        # Revisar ambos archivos
        files = [
            "EXCEL-V12-FINAL.xlsx",
            "template-picks-and-tipsters.xlsx"
        ]
        for f in files:
            try:
                check_dashboard_references(f)
            except FileNotFoundError:
                print(f"❌ Archivo no encontrado: {f}")
