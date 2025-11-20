#!/usr/bin/env python3
"""
🔍 INSPECTOR DE DASHBOARDS EXPORTADOS

Script para diagnosticar por qué los dashboards no muestran datos dinámicos.
Inspecciona:
1. Si las celdas tienen fórmulas o valores estáticos
2. Si las fórmulas son correctas
3. Si las referencias de hojas son correctas
"""

import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def inspect_dashboard(filepath, sheet_name):
    """
    Inspecciona un dashboard específico
    """
    print(f"\n{'='*80}")
    print(f"🔍 INSPECCIONANDO: {sheet_name}")
    print('='*80)
    
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    
    # Inspeccionar fila 3 (primer tipster)
    print("\n📋 FILA 3 (Primer Tipster):")
    print("-" * 80)
    
    # Columnas importantes para verificar
    columns_to_check = {
        'A': 'TIPSTER (nombre)',
        'B': 'Juega a Unidades',
        'C': 'Beneficio UDS (SUMIF)',
        'D': 'Beneficio € (SUMIF)',
        'E': 'Apostado (SUMIF)',
        'F': 'YIELD (fórmula)',
        'G': 'Tips totales (COUNTIFS)',
        'H': 'Tips W (COUNTIFS)',
        'I': 'Tips L (COUNTIFS)',
        'J': 'Tips V (COUNTIFS)',
        'K': '% Aciertos (fórmula)',
        'N': 'TENIS % (COUNTIFS sport)',
        'O': 'BALONCESTO % (COUNTIFS sport)'
    }
    
    for col, description in columns_to_check.items():
        cell = ws[f'{col}3']
        cell_value = cell.value
        cell_type = type(cell_value).__name__
        
        # Verificar si tiene fórmula
        if hasattr(cell, '_value') and isinstance(cell._value, str) and cell._value.startswith('='):
            formula = cell._value
            print(f"\n✅ {col}3: {description}")
            print(f"   Tipo: FÓRMULA")
            print(f"   Fórmula: {formula}")
        elif cell.data_type == 'f':
            # Otra forma de detectar fórmulas
            print(f"\n✅ {col}3: {description}")
            print(f"   Tipo: FÓRMULA (data_type=f)")
            print(f"   Valor calculado: {cell_value}")
        else:
            print(f"\n❌ {col}3: {description}")
            print(f"   Tipo: {cell_type}")
            print(f"   Valor: {cell_value}")
            print(f"   ⚠️  PROBLEMA: No es una fórmula, es un valor estático")
    
    # Verificar fila 2 (headers)
    print("\n\n📋 FILA 2 (Headers de deportes):")
    print("-" * 80)
    sport_cols = ['N', 'O', 'P', 'Q', 'R', 'S', 'T']
    for col in sport_cols:
        cell = ws[f'{col}2']
        print(f"{col}2: {cell.value}")
    
    wb.close()

def main():
    if len(sys.argv) < 2:
        print("❌ Uso: python3 inspect-exported-dashboard.py <archivo.xlsx>")
        sys.exit(1)
    
    filepath = sys.argv[1]
    
    print(f"\n🔍 INSPECTOR DE DASHBOARDS")
    print(f"📁 Archivo: {filepath}")
    
    try:
        # Inspeccionar ambos dashboards
        inspect_dashboard(filepath, 'Mis_Picks_Dashboard')
        inspect_dashboard(filepath, 'Tipster_Picks_Dashboard')
        
        print("\n" + "="*80)
        print("✅ INSPECCIÓN COMPLETADA")
        print("="*80)
        
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
