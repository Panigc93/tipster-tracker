#!/usr/bin/env python3
"""
Analizar las fórmulas del dashboard para ver cómo se conectan con Realizadas
"""

import openpyxl
import sys

def analyze_dashboard_formulas(filename):
    print(f"\n📊 Analizando fórmulas de dashboards en: {filename}")
    print("=" * 80)
    
    wb = openpyxl.load_workbook(filename, data_only=False)
    
    # Analizar Mis_Picks_Dashboard
    if 'Mis_Picks_Dashboard' in wb.sheetnames:
        ws = wb['Mis_Picks_Dashboard']
        print(f"\n📋 HOJA: Mis_Picks_Dashboard")
        print("-" * 80)
        
        # Ver fila 3 (primera fila de datos después de headers)
        print("\n🔍 FILA 3 (primera fila de datos):\n")
        
        # Revisar cada columna de la fila 3
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
            cell = ws[f'{col}3']
            col_letter = col
            header = ws[f'{col}2'].value if ws[f'{col}2'].value else ''
            
            value = cell.value
            
            if value and isinstance(value, str) and value.startswith('='):
                # Es una fórmula
                print(f"   {col_letter}3 ({header}): FÓRMULA")
                print(f"      {value}")
            else:
                # Es un valor fijo
                print(f"   {col_letter}3 ({header}): VALOR = {value}")
    
    # Analizar Tipster_Picks_Dashboard
    if 'Tipster_Picks_Dashboard' in wb.sheetnames:
        ws = wb['Tipster_Picks_Dashboard']
        print(f"\n\n📋 HOJA: Tipster_Picks_Dashboard")
        print("-" * 80)
        
        print("\n🔍 FILA 3 (primera fila de datos):\n")
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
            cell = ws[f'{col}3']
            col_letter = col
            header = ws[f'{col}2'].value if ws[f'{col}2'].value else ''
            
            value = cell.value
            
            if value and isinstance(value, str) and value.startswith('='):
                print(f"   {col_letter}3 ({header}): FÓRMULA")
                print(f"      {value}")
            else:
                print(f"   {col_letter}3 ({header}): VALOR = {value}")
    
    wb.close()
    print("\n" + "=" * 80 + "\n")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 analyze-dashboard-formulas.py <archivo.xlsx>")
        sys.exit(1)
    
    analyze_dashboard_formulas(sys.argv[1])
