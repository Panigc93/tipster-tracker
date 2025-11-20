#!/usr/bin/env python3
"""
Script de prueba: verifica que el Excel generado tenga las referencias
correctas a los dashboards en las fórmulas VLOOKUP.

Expectativa:
- Realizadas → debe apuntar a Mis_Picks_Dashboard
- Lanzadas Tipster → debe apuntar a Tipster_Picks_Dashboard
"""

import openpyxl
import sys

def verify_excel_dashboard_refs(filename):
    print(f"\n🔍 VERIFICACIÓN DE REFERENCIAS A DASHBOARDS")
    print("=" * 80)
    print(f"📊 Archivo: {filename}\n")
    
    wb = openpyxl.load_workbook(filename, data_only=False)
    
    expected_refs = {
        'Realizadas': 'Mis_Picks_Dashboard',
        'Lanzadas Tipster': 'Tipster_Picks_Dashboard'
    }
    
    all_correct = True
    
    for sheet_name, expected_dashboard in expected_refs.items():
        if sheet_name not in wb.sheetnames:
            print(f"❌ Sheet '{sheet_name}' no encontrada")
            all_correct = False
            continue
            
        ws = wb[sheet_name]
        print(f"📋 SHEET: '{sheet_name}'")
        print("-" * 80)
        
        # Revisar fila 7, columna G (primera fila de datos con VLOOKUP)
        cell_g7 = ws['G7']
        
        if cell_g7.value and isinstance(cell_g7.value, str) and 'VLOOKUP' in cell_g7.value:
            formula = cell_g7.value
            
            # Verificar si contiene el dashboard esperado
            if expected_dashboard in formula:
                print(f"✅ CORRECTO: Apunta a '{expected_dashboard}'")
                print(f"   Fórmula: {formula[:100]}...")
            else:
                print(f"❌ ERROR: NO apunta a '{expected_dashboard}'")
                print(f"   Fórmula encontrada: {formula}")
                all_correct = False
                
                # Detectar a qué dashboard apunta
                if 'Mis_Picks_Dashboard' in formula:
                    print(f"   ⚠️  Apunta a: Mis_Picks_Dashboard (incorrecto)")
                elif 'Tipster_Picks_Dashboard' in formula:
                    print(f"   ⚠️  Apunta a: Tipster_Picks_Dashboard (incorrecto)")
        else:
            print(f"❌ No se encontró VLOOKUP en G7")
            print(f"   Valor de G7: {cell_g7.value}")
            all_correct = False
        
        print()
    
    wb.close()
    
    print("=" * 80)
    if all_correct:
        print("✅ VERIFICACIÓN EXITOSA: Todas las referencias son correctas")
    else:
        print("❌ VERIFICACIÓN FALLIDA: Hay referencias incorrectas")
    print("=" * 80 + "\n")
    
    return all_correct

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 verify-dashboard-refs.py <archivo.xlsx>")
        sys.exit(1)
    
    filename = sys.argv[1]
    success = verify_excel_dashboard_refs(filename)
    sys.exit(0 if success else 1)
